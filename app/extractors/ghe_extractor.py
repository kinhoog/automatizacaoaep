"""Official GHE spreadsheet extractor.

The extractor deliberately has no field for employee names. If a workbook
represents population as one person per row, only cell presence is counted;
the individual value is never retained, returned or logged.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models import GHE, GHEExtractionResult
from app.services.normalization import (
    canonical_ghe_code,
    clean_text,
    normalize_key,
    parse_ghe_reference,
    parse_population,
    stable_unique,
)

from . import ExtractionError


@dataclass(slots=True)
class _Columns:
    code: int | None = None
    name: int | None = None
    combined: int | None = None
    sector: int | None = None
    role: int | None = None
    population: int | None = None
    person: list[int] = field(default_factory=list)

    @property
    def score(self) -> int:
        identity = 3 if self.combined is not None else 0
        identity += 2 if self.code is not None else 0
        identity += 1 if self.name is not None else 0
        detail = sum(
            value is not None
            for value in (self.sector, self.role, self.population)
        )
        if self.person:
            detail += 1
        return identity + detail

    @property
    def usable(self) -> bool:
        has_identity = self.combined is not None or self.code is not None
        return has_identity and self.score >= 4


@dataclass(slots=True)
class _Group:
    code: str
    name: str
    rows: list[int] = field(default_factory=list)
    sectors: list[str] = field(default_factory=list)
    roles: list[str] = field(default_factory=list)
    population_values: list[int] = field(default_factory=list)
    person_row_count: int = 0


def _column_kind(value: Any) -> str | None:
    key = normalize_key(value)
    if not key:
        return None
    if any(
        phrase in key
        for phrase in (
            "nome colaborador",
            "nome funcionario",
            "nome empregado",
            "colaborador nome",
            "funcionario nome",
            "trabalhador nome",
            "matricula",
            "cpf",
        )
    ):
        return "person"
    if (
        key in {"ghe", "identificacao ghe", "grupo homogeneo de exposicao"}
        or ("ghe" in key and any(word in key for word in ("identificacao", "grupo")))
    ):
        return "combined"
    if "ghe" in key and any(
        word in key for word in ("codigo", "cod", "numero", "nro", "id")
    ):
        return "code"
    if "ghe" in key and any(
        word in key for word in ("nome", "descricao", "denominacao")
    ):
        return "name"
    if key in {"codigo ghe", "cod ghe", "numero ghe"}:
        return "code"
    if any(word in key for word in ("setor", "departamento", "area", "unidade")):
        return "sector"
    if any(word in key for word in ("cargo", "funcao", "ocupacao", "posto")):
        return "role"
    if (
        key in {"qtd", "qtde", "quantidade", "populacao", "headcount"}
        or any(
            phrase in key
            for phrase in (
                "quantidade colaboradores",
                "numero colaboradores",
                "total colaboradores",
                "qtd colaboradores",
                "numero trabalhadores",
                "total pessoas",
            )
        )
    ):
        return "population"
    return None


def _map_header(values: tuple[Any, ...]) -> _Columns:
    columns = _Columns()
    for index, value in enumerate(values, start=1):
        kind = _column_kind(value)
        if kind == "person":
            columns.person.append(index)
        elif kind and getattr(columns, kind) is None:
            setattr(columns, kind, index)
    return columns


def _ensure_worksheet_dimensions(worksheet: Worksheet) -> None:
    """Populate missing bounds in valid XLSX files without a dimension tag."""

    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.calculate_dimension(force=True)


def _find_header(workbook: Any) -> tuple[Worksheet, int, _Columns]:
    best: tuple[int, Worksheet, int, _Columns] | None = None
    for worksheet in workbook.worksheets:
        _ensure_worksheet_dimensions(worksheet)
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(60, worksheet.max_row)),
            start=1,
        ):
            columns = _map_header(tuple(cell.value for cell in row))
            candidate = (columns.score, worksheet, row_number, columns)
            if best is None or candidate[0] > best[0]:
                best = candidate
            if columns.usable and columns.score >= 7:
                return worksheet, row_number, columns
    if best and best[3].usable:
        return best[1], best[2], best[3]
    raise ExtractionError(
        "Não foi possível identificar o cabeçalho oficial de GHEs na planilha."
    )


def _cell(row: tuple[Any, ...], one_based_index: int | None) -> Any:
    if one_based_index is None or one_based_index > len(row):
        return None
    return row[one_based_index - 1]


def _coerce_separate_code(value: Any) -> str | None:
    canonical = canonical_ghe_code(value)
    if canonical:
        return canonical
    text = clean_text(value)
    if text.isdigit():
        return f"GHE {int(text):02d}"
    if isinstance(value, (int, float)) and float(value).is_integer():
        return f"GHE {int(value):02d}"
    return None


def _split_values(value: Any) -> list[str]:
    return [
        clean_text(item)
        for item in re.split(r"[;|\r\n]+", clean_text(value))
        if clean_text(item)
    ]


class GHEExtractor:
    """Read an XLSX workbook in read-only/data-only mode."""

    def extract(self, source: str | Path) -> GHEExtractionResult:
        source_path = Path(source)
        try:
            workbook = load_workbook(
                filename=source_path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise ExtractionError(
                "A planilha de GHEs não é um XLSX válido ou está corrompida."
            ) from exc

        try:
            worksheet, header_row, columns = _find_header(workbook)
            result = self._extract_rows(worksheet, header_row, columns)
            self._enrich_from_detail_sheets(workbook, result)
            return result
        finally:
            workbook.close()

    def _enrich_from_detail_sheets(
        self,
        workbook: Any,
        result: GHEExtractionResult,
    ) -> None:
        """Add official sectors/roles from per-GHE sheets without person data."""

        by_code = {ghe.canonical_code: ghe for ghe in result.ghes}
        found_details: set[str] = set()
        ignored_headers = list(result.ignored_person_columns)
        for worksheet in workbook.worksheets:
            _ensure_worksheet_dimensions(worksheet)
            sheet_code = canonical_ghe_code(worksheet.title)
            if sheet_code not in by_code:
                continue
            header_row: int | None = None
            columns: _Columns | None = None
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(40, worksheet.max_row),
                    values_only=True,
                ),
                start=1,
            ):
                mapped = _map_header(tuple(row))
                if mapped.sector is not None and mapped.role is not None:
                    header_row = row_number
                    columns = mapped
                    for person_column in mapped.person:
                        header = clean_text(_cell(tuple(row), person_column))
                        if header:
                            ignored_headers.append(header)
                    break
            if header_row is None or columns is None:
                continue
            sectors: list[str] = []
            roles: list[str] = []
            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                max_row=worksheet.max_row,
                values_only=True,
            ):
                sectors.extend(_split_values(_cell(row, columns.sector)))
                roles.extend(_split_values(_cell(row, columns.role)))
            if sectors or roles:
                ghe = by_code[sheet_code]
                ghe.sectors = stable_unique([*ghe.sectors, *sectors])
                ghe.roles = stable_unique([*ghe.roles, *roles])
                found_details.add(sheet_code)
        result.ignored_person_columns = stable_unique(ignored_headers)
        missing = sorted(set(by_code) - found_details)
        for code in missing:
            result.warnings.append(
                f"{code}: setores/cargos detalhados não foram localizados em aba própria."
            )

    def _extract_rows(
        self, worksheet: Worksheet, header_row: int, columns: _Columns
    ) -> GHEExtractionResult:
        groups: dict[str, _Group] = {}
        order: list[str] = []
        active_code: str | None = None
        warnings: list[str] = []
        explicit_total: int | None = None

        rows = worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            values_only=True,
        )
        for row_number, row in enumerate(rows, start=header_row + 1):
            visible_values = [
                clean_text(_cell(row, index))
                for index in (
                    columns.code,
                    columns.name,
                    columns.combined,
                    columns.sector,
                    columns.role,
                    columns.population,
                )
                if index is not None
            ]
            if not any(visible_values) and not any(
                bool(_cell(row, index)) for index in columns.person
            ):
                continue

            identity_value = _cell(row, columns.combined)
            code_value = _cell(row, columns.code)
            name_value = _cell(row, columns.name)
            identity_text = clean_text(identity_value)
            code, parsed_name = parse_ghe_reference(identity_text)
            if not code:
                code = _coerce_separate_code(code_value)
            row_key = normalize_key(" ".join(visible_values))
            is_total_row = row_key.startswith(("total ", "total geral", "subtotal "))
            population_value = parse_population(_cell(row, columns.population))
            if is_total_row:
                if population_value is not None:
                    explicit_total = population_value
                continue

            if code:
                active_code = code
                name = clean_text(name_value) or parsed_name or ""
                if code not in groups:
                    groups[code] = _Group(code=code, name=name)
                    order.append(code)
                elif name and not groups[code].name:
                    groups[code].name = name
            elif active_code:
                code = active_code
            else:
                # Introductory or explanatory rows outside the official table
                # do not become synthetic GHEs.
                continue

            group = groups[code]
            group.rows.append(row_number)
            group.sectors.extend(_split_values(_cell(row, columns.sector)))
            group.roles.extend(_split_values(_cell(row, columns.role)))
            if population_value is not None:
                group.population_values.append(population_value)
            elif columns.population is None and columns.person:
                # Presence-only check: the personal value is never retained.
                if any(
                    clean_text(_cell(row, person_column))
                    for person_column in columns.person
                ):
                    group.person_row_count += 1

        if not groups:
            raise ExtractionError("Nenhum GHE oficial foi encontrado na planilha.")

        ghes: list[GHE] = []
        used_person_count = False
        for code in order:
            group = groups[code]
            if not group.name:
                warnings.append(f"{code}: nome oficial ausente na planilha.")
            if group.population_values:
                population = sum(group.population_values)
            else:
                population = group.person_row_count
                used_person_count = used_person_count or population > 0
            if population == 0:
                warnings.append(f"{code}: população oficial igual a zero ou ausente.")
            ghes.append(
                GHE(
                    code=group.code,
                    name=group.name or group.code,
                    sectors=stable_unique(group.sectors),
                    roles=stable_unique(group.roles),
                    population=population,
                    source_rows=group.rows,
                )
            )

        total = sum(ghe.population for ghe in ghes)
        if explicit_total is not None and explicit_total != total:
            warnings.append(
                "O total geral declarado na planilha diverge da soma dos GHEs."
            )
        if used_person_count:
            warnings.append(
                "A população foi contada por presença de linhas individuais; "
                "nenhum nome foi armazenado."
            )

        ignored_headers: list[str] = []
        header_values = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )
        for column in columns.person:
            header = clean_text(_cell(header_values, column))
            if header:
                ignored_headers.append(header)

        return GHEExtractionResult(
            ghes=ghes,
            source_sheet=worksheet.title,
            header_row=header_row,
            ignored_person_columns=stable_unique(ignored_headers),
            warnings=warnings,
        )


def extract_ghes(source: str | Path) -> GHEExtractionResult:
    """Functional API for dependency-light callers."""

    return GHEExtractor().extract(source)
